"""
OncoSignal Free Pipeline v2
============================
Pulls REAL global oncology company data from 7 free sources.
No API keys. No paid databases. Catches pharma companies at EVERY stage.

Sources:
  1. ClinicalTrials.gov  — US trial sponsors, phase, cancer type
  2. WHO ICTRP           — GLOBAL registry (catches ASX/EU/Asian companies)
  3. PubMed / NCBI       — Companies publishing oncology research
  4. bioRxiv / medRxiv   — Preprints (catches pre-clinical companies early)
  5. OpenFDA             — US drug application sponsors
  6. SEC EDGAR           — Recent oncology IPO / S-1 filings
  7. Google News RSS     — Funding announcements, stealth launches, Series A/B
                           THIS is the key source for catching new pharma companies
                           like Radiopharm Theranostics BEFORE they file trials

Why each source catches a different company type:
  -> Google News RSS : Series A/B raises, stealth launches — EARLIEST signal
  -> WHO ICTRP       : ASX/LSE/TSX-listed companies with non-US trials
  -> PubMed          : Companies publishing before they file trials
  -> bioRxiv         : Pre-clinical stage, no IND yet
  -> ClinicalTrials  : US-registered active trials
  -> EDGAR           : US IPOs/S-1 filings

Install:  pip install requests openpyxl
Run:      python oncology_pipeline_v2.py
          python oncology_pipeline_v2.py --sector radiopharmaceutical
          python oncology_pipeline_v2.py --sector all --since 2021 --count 40
          python oncology_pipeline_v2.py --sources gnews,who --sector radiopharmaceutical
"""

import requests, re, time, argparse, json, xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from collections import defaultdict

# ── CONFIG ────────────────────────────────────────────────────────────────────

HEADERS = {"User-Agent": "OncoSignal-Pipeline/2.0 research@oncosignal.io"}

EXCLUDE_ORGS = [
    "roche","pfizer","novartis","astrazeneca","bristol-myers","merck",
    "johnson & johnson","eli lilly","abbvie","amgen","genentech","bayer",
    "sanofi","glaxosmithkline","boehringer","nih","nci","national cancer",
    "national institute","university","hospital","medical center","medical school",
    "health system","clinic","institute of","foundation","mayo clinic",
    "memorial sloan","md anderson","fred hutchinson","dana-farber",
    "johns hopkins","stanford","harvard","yale","columbia","michigan",
    "nyu langone","ucsf","ucla","government","department of","ministry of",
    "centre for","center for","children's","veterans","kaiser",
]

SECTOR_KEYWORDS = {
    "radiopharmaceutical": [
        "radiopharmaceutical","theranostic","radioligand","lutetium-177","actinium-225",
        "PSMA","radioconjugate","alpha emitter","beta emitter","targeted radionuclide",
        "radiolabeled","radioimmunotherapy","Lu-177","Ac-225","DOTATATE","DOTATOC",
    ],
    "immuno-oncology": [
        "immunotherapy","checkpoint inhibitor","PD-1","PD-L1","CTLA-4",
        "CAR-T","NK cell","TIL therapy","bispecific antibody","immune checkpoint",
        "tumor infiltrating","cancer vaccine","oncolytic virus",
    ],
    "adc": [
        "antibody-drug conjugate","ADC","drug conjugate","payload linker",
        "HER2 ADC","TROP2","nectin-4","folate receptor",
    ],
    "precision": [
        "EGFR","KRAS","BRAF V600","HER2","ALK inhibitor","RET fusion",
        "MET amplification","FGFR","precision oncology","targeted therapy",
        "biomarker-driven","tumor mutational burden","microsatellite instability",
    ],
    "cell therapy": [
        "CAR-T cell","cell therapy","T-cell engager","NK cell therapy",
        "TIL therapy","stem cell transplant","allogeneic","autologous",
    ],
    "rna": [
        "mRNA oncology","siRNA cancer","RNA interference","circular RNA",
        "antisense oligonucleotide","LNP oncology","lipid nanoparticle cancer",
    ],
    "all": [
        "oncology","cancer","tumor","carcinoma","lymphoma",
        "leukemia","sarcoma","melanoma","glioblastoma",
    ],
}

# ══════════════════════════════════════════════════════════════════════════════
# SOURCE 1 — ClinicalTrials.gov (US registry)
# https://clinicaltrials.gov/data-api/api
# ══════════════════════════════════════════════════════════════════════════════

def fetch_clinicaltrials(sector="all", since_year=2020, max_results=100):
    print(f"\n[1/6] ClinicalTrials.gov (US) — since {since_year}...")
    terms     = SECTOR_KEYWORDS.get(sector, SECTOR_KEYWORDS["all"])
    query     = " OR ".join(terms[:5])
    url       = "https://clinicaltrials.gov/api/v2/studies"
    params    = {
        "query.cond":           "cancer OR oncology OR tumor",
        "query.term":           query,
        "filter.advanced":      f"AREA[StartDate]RANGE[{since_year}-01-01,MAX]",
        "filter.overallStatus": "RECRUITING,ACTIVE_NOT_RECRUITING,NOT_YET_RECRUITING",
        "fields": ",".join([
            "protocolSection.identificationModule.nctId",
            "protocolSection.identificationModule.briefTitle",
            "protocolSection.sponsorCollaboratorsModule.leadSponsor",
            "protocolSection.statusModule.startDateStruct",
            "protocolSection.statusModule.overallStatus",
            "protocolSection.conditionsModule.conditions",
            "protocolSection.designModule.phases",
        ]),
        "pageSize": min(max_results, 100),
        "format":   "json",
        "sort":     "StartDate:desc",
    }
    try:
        r = requests.get(url, params=params, headers=HEADERS, timeout=20)
        r.raise_for_status()
        studies = r.json().get("studies", [])
        result  = _parse_ct(studies)
        print(f"    {len(studies)} trials → {len(result)} sponsors")
        return result
    except Exception as e:
        print(f"    ERROR: {e}")
        return []

def _parse_ct(studies):
    bucket = defaultdict(lambda: {
        "source":"ClinicalTrials.gov (US)","trials":[],"phases":[],
        "conditions":[],"latest_start":"","region":"USA",
    })
    for s in studies:
        p       = s.get("protocolSection", {})
        sponsor = p.get("sponsorCollaboratorsModule", {}).get("leadSponsor", {})
        name    = sponsor.get("name","").strip()
        if not name or sponsor.get("class","") in ("NIH","U.S. Fed") or _is_large(name):
            continue
        start  = p.get("statusModule",{}).get("startDateStruct",{}).get("date","")
        phases = p.get("designModule",{}).get("phases",[])
        conds  = p.get("conditionsModule",{}).get("conditions",[])
        c = bucket[name]
        c["name"] = name
        c["trials"].append(p.get("identificationModule",{}).get("briefTitle",""))
        c["phases"].extend(phases)
        c["conditions"].extend(conds[:2])
        if start and start > c["latest_start"]:
            c["latest_start"] = start
    return list(bucket.values())


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE 2 — WHO ICTRP (GLOBAL registry — catches AU, EU, Asia companies)
# https://trialsearch.who.int  — XML RSS feed, no key needed
# This is the key source for catching ASX-listed companies like Radiopharm
# ══════════════════════════════════════════════════════════════════════════════

def fetch_who_ictrp(sector="all", since_year=2020):
    print(f"\n[2/6] WHO ICTRP (Global registry) — catches AU/EU/Asian companies...")
    terms   = SECTOR_KEYWORDS.get(sector, SECTOR_KEYWORDS["all"])
    results = []

    for term in terms[:4]:
        # WHO ICTRP public search — returns XML
        url = "https://trialsearch.who.int/API/json"
        params = {
            "action":   "trial",
            "searchterms": term,
            "country":  "",     # blank = all countries
            "status":   "pending,recruiting,ongoing",
            "phase":    "",
            "study_type": "interventional",
            "daterange": f"{since_year}-01-01 to {date.today().isoformat()}",
        }
        try:
            r = requests.get(url, params=params, headers=HEADERS, timeout=20)
            if r.status_code == 200:
                data = r.json() if r.headers.get("content-type","").startswith("application/json") else {}
                trials = data.get("Trials", data.get("trials", []))
                results.extend(_parse_who(trials, term))
            time.sleep(0.8)
        except Exception as e:
            # Fallback: try the RSS/XML endpoint
            try:
                rss_url = f"https://trialsearch.who.int/Trial2.aspx?TrialID=&Title={requests.utils.quote(term)}&InternationalId=&ReferenceNumber=&WhoCoDomain=&EudraCTNumber=&IssuerName=&StudyType=interventional&Phase=&RecruitingStatus=recruiting&Gender=&PatientAgeGroup=&StudyDesign=&PrimarySponsors=&Conditions=oncology+{requests.utils.quote(term)}&Interventions=&Countries=&StartDate=&DateRange=&OutputFormat=RSS&Register=All&action=search"
                r2 = requests.get(rss_url, headers=HEADERS, timeout=20)
                if r2.status_code == 200:
                    results.extend(_parse_who_rss(r2.text, term))
            except Exception:
                pass

    unique = _dedup_by_name(results)
    print(f"    {len(unique)} global sponsors found")
    return unique

def _parse_who(trials, term):
    companies = []
    for t in trials if isinstance(trials, list) else []:
        sponsor = (t.get("primary_sponsor") or t.get("PrimarySponsors") or "").strip()
        if not sponsor or _is_large(sponsor):
            continue
        country = t.get("countries") or t.get("country") or ""
        phase   = t.get("phase") or t.get("Phase") or ""
        cond    = t.get("condition") or t.get("Conditions") or "Oncology"
        date_r  = t.get("date_registration") or t.get("DateRegistration") or ""
        companies.append({
            "name":         sponsor,
            "source":       "WHO ICTRP (Global)",
            "region":       str(country)[:30],
            "phases":       [phase] if phase else [],
            "conditions":   [str(cond)[:40]],
            "trials":       [t.get("scientific_title","")[:80] or term],
            "latest_start": _clean_date(str(date_r)),
        })
    return companies

def _parse_who_rss(xml_text, term):
    companies = []
    try:
        root = ET.fromstring(xml_text)
        ns   = {"dc": "http://purl.org/dc/elements/1.1/"}
        for item in root.iter("item"):
            title    = item.findtext("title","")
            desc     = item.findtext("description","")
            sponsor  = item.findtext("dc:creator", namespaces=ns) or _extract_sponsor_from_text(desc)
            if not sponsor or _is_large(sponsor):
                continue
            pub_date = item.findtext("pubDate","")
            companies.append({
                "name":         sponsor.strip(),
                "source":       "WHO ICTRP (Global)",
                "region":       "Global",
                "phases":       _extract_phase(title + " " + desc),
                "conditions":   _extract_conditions(title + " " + desc),
                "trials":       [title[:80]],
                "latest_start": _clean_date(pub_date),
            })
    except Exception:
        pass
    return companies

def _extract_sponsor_from_text(text):
    m = re.search(r"(?:sponsor|sponsored by|primary sponsor)[:\s]+([A-Z][A-Za-z0-9 ,\.]{3,40})", text, re.I)
    return m.group(1).strip() if m else ""


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE 3 — PubMed / NCBI E-utilities (free, no key needed up to 10 req/sec)
# https://www.ncbi.nlm.nih.gov/books/NBK25500/
# Finds companies publishing oncology research — catches pre-IND companies
# ══════════════════════════════════════════════════════════════════════════════

def fetch_pubmed(sector="all", since_year=2020, max_results=100):
    print(f"\n[3/6] PubMed / NCBI — companies publishing oncology research...")
    terms    = SECTOR_KEYWORDS.get(sector, SECTOR_KEYWORDS["all"])
    query    = " OR ".join(f'"{t}"' for t in terms[:4])
    # Filter: affiliation contains biotech/pharma/therapeutics keywords
    full_q   = (f'({query})[Title/Abstract] AND '
                f'("therapeutics"[Affiliation] OR "pharmaceuticals"[Affiliation] OR '
                f'"biosciences"[Affiliation] OR "biopharma"[Affiliation] OR '
                f'"oncology inc"[Affiliation]) AND '
                f'{since_year}/01/01:{date.today().year}/12/31[Date - Publication]')

    # Step 1: search for PMIDs
    search_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
    params     = {
        "db":      "pubmed",
        "term":    full_q,
        "retmax":  max_results,
        "retmode": "json",
        "sort":    "pub_date",
    }
    try:
        r = requests.get(search_url, params=params, headers=HEADERS, timeout=20)
        r.raise_for_status()
        pmids = r.json().get("esearchresult", {}).get("idlist", [])
        if not pmids:
            print(f"    0 articles found")
            return []
        print(f"    {len(pmids)} articles found — fetching affiliations...")
        time.sleep(0.4)

        # Step 2: fetch summaries to extract affiliations (company names)
        fetch_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
        params2   = {
            "db":      "pubmed",
            "id":      ",".join(pmids[:50]),  # batch of 50
            "retmode": "xml",
            "rettype": "abstract",
        }
        r2 = requests.get(fetch_url, params=params2, headers=HEADERS, timeout=30)
        r2.raise_for_status()
        result = _parse_pubmed_xml(r2.text)
        print(f"    {len(result)} companies extracted from affiliations")
        return result
    except Exception as e:
        print(f"    ERROR: {e}")
        return []

def _parse_pubmed_xml(xml_text):
    bucket = defaultdict(lambda: {
        "source":"PubMed (NCBI)","trials":[],"phases":[],
        "conditions":[],"latest_start":"","region":"",
    })
    try:
        root = ET.fromstring(xml_text)
        for article in root.iter("PubmedArticle"):
            # Publication date
            pub_date = ""
            for pd in article.iter("PubDate"):
                yr  = pd.findtext("Year","")
                mon = pd.findtext("Month","01")
                if yr:
                    pub_date = f"{yr}-{mon[:2] if mon.isdigit() else '01'}-01"
                    break

            # Title
            title = article.findtext(".//ArticleTitle","")

            # Affiliations — look for company names
            for aff in article.iter("Affiliation"):
                text  = aff.text or ""
                cname = _company_from_affiliation(text)
                if not cname or _is_large(cname):
                    continue
                country = _country_from_affiliation(text)
                c = bucket[cname]
                c["name"]       = cname
                c["region"]     = country
                c["trials"].append(title[:80])
                c["conditions"].extend(_extract_conditions(title))
                c["phases"].extend(_extract_phase(title))
                if pub_date and pub_date > c["latest_start"]:
                    c["latest_start"] = pub_date
    except ET.ParseError:
        pass
    return list(bucket.values())

def _company_from_affiliation(text):
    """Extract company name from affiliation string like:
    'Discovery Biology, Acme Therapeutics, San Francisco, CA 94105, USA'"""
    # Match patterns ending in biotech suffixes before a comma/period
    patterns = [
        r"([A-Z][A-Za-z0-9\-]+(?: [A-Z][A-Za-z0-9\-]+){0,4})\s*(?:Inc\.?|Corp\.?|Ltd\.?|LLC|GmbH|AG|SA|SAS|BV|NV),",
        r"([A-Z][A-Za-z0-9\-]+(?: [A-Z][A-Za-z0-9\-]+){0,4} (?:Therapeutics|Pharmaceuticals|Biosciences|Bioscience|Biotech|Biologics|Oncology|Biopharma|Bio|Medicines))[,\.]",
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if m:
            name = m.group(1).strip()
            if len(name) > 4:
                return name
    return ""

def _country_from_affiliation(text):
    countries = {
        "USA":"USA","United States":"USA","U.S.A":"USA",
        "Australia":"Australia","UK":"UK","United Kingdom":"UK",
        "Germany":"Germany","France":"France","Switzerland":"Switzerland",
        "Canada":"Canada","Japan":"Japan","China":"China",
        "Netherlands":"Netherlands","Belgium":"Belgium","Sweden":"Sweden",
        "Israel":"Israel","Denmark":"Denmark","South Korea":"South Korea",
    }
    for k, v in countries.items():
        if k in text:
            return v
    return ""


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE 4 — bioRxiv + medRxiv (preprints — catches pre-clinical companies)
# https://api.biorxiv.org  — free, no key
# Companies publishing preprints = very early stage, no data vendor yet
# ══════════════════════════════════════════════════════════════════════════════

def fetch_biorxiv(sector="all", since_year=2020):
    print(f"\n[4/6] bioRxiv / medRxiv — pre-clinical preprints...")
    terms   = SECTOR_KEYWORDS.get(sector, SECTOR_KEYWORDS["all"])
    results = []
    start   = f"{since_year}-01-01"
    end     = date.today().isoformat()

    # bioRxiv API: /details/{server}/{date_start}/{date_end}/{cursor}
    for server in ["biorxiv", "medrxiv"]:
        cursor = 0
        while cursor < 60:  # max 3 pages of 20
            url = f"https://api.biorxiv.org/details/{server}/{start}/{end}/{cursor}/json"
            try:
                r = requests.get(url, headers=HEADERS, timeout=20)
                r.raise_for_status()
                data       = r.json()
                collection = data.get("collection", [])
                if not collection:
                    break

                # Filter to sector keywords
                for paper in collection:
                    abstract = (paper.get("abstract","") + " " + paper.get("title","")).lower()
                    if not any(t.lower() in abstract for t in terms):
                        continue
                    author_aff = paper.get("author_corresponding_institution","")
                    cname = _company_from_affiliation(author_aff)
                    if not cname or _is_large(cname):
                        # try extracting from abstract text
                        cname = _extract_company_mention(abstract)
                    if not cname or _is_large(cname):
                        continue

                    pub_date = _clean_date(paper.get("date",""))
                    results.append({
                        "name":         cname,
                        "source":       f"{'bioRxiv' if server=='biorxiv' else 'medRxiv'} (Preprint)",
                        "region":       _country_from_affiliation(author_aff),
                        "phases":       _extract_phase(paper.get("title","") + " " + paper.get("abstract","")),
                        "conditions":   _extract_conditions(paper.get("title","") + " " + paper.get("abstract","")),
                        "trials":       [paper.get("title","")[:80]],
                        "latest_start": pub_date,
                    })
                cursor += 20
                time.sleep(0.5)
            except Exception as e:
                print(f"    {server} page {cursor}: {e!s:.50}")
                break

    unique = _dedup_by_name(results)
    print(f"    {len(unique)} companies from preprints")
    return unique

def _extract_company_mention(text):
    """Find company names mentioned in paper abstracts."""
    m = re.search(
        r"\b([A-Z][A-Za-z0-9]+(?: [A-Z][A-Za-z0-9]+){0,3} "
        r"(?:Therapeutics|Pharmaceuticals|Biosciences|Biotech|Oncology|Biopharma|Bio|Medicines))\b",
        text, re.I
    )
    return m.group(1).strip() if m else ""


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE 5 — OpenFDA drug applications
# https://open.fda.gov/apis/drug/drugsfda/  — free, no key
# ══════════════════════════════════════════════════════════════════════════════

def fetch_openfda(sector="all", since_year=2020):
    print(f"\n[5/6] OpenFDA — drug application sponsors...")
    terms   = SECTOR_KEYWORDS.get(sector, SECTOR_KEYWORDS["all"])
    results = []

    for term in terms[:3]:
        url    = "https://api.fda.gov/drug/drugsfda.json"
        params = {
            "search": f'products.brand_name:"{term}"',
            "limit":  50,
        }
        try:
            r = requests.get(url, params=params, headers=HEADERS, timeout=15)
            if r.status_code == 200:
                for rec in r.json().get("results", []):
                    sponsor = rec.get("sponsor_name","").strip()
                    if not sponsor or _is_large(sponsor):
                        continue
                    subs     = rec.get("submissions",[])
                    last_sub = max((s.get("submission_status_date","") for s in subs), default="")
                    yr       = int(last_sub[:4]) if last_sub and last_sub[:4].isdigit() else 0
                    if yr < since_year:
                        continue
                    results.append({
                        "name":         sponsor,
                        "source":       "OpenFDA",
                        "region":       "USA",
                        "phases":       ["NDA/BLA Filed"],
                        "conditions":   ["Oncology"],
                        "trials":       [rec.get("application_number","")],
                        "latest_start": _clean_date(last_sub),
                    })
            time.sleep(0.4)
        except Exception as e:
            print(f"    OpenFDA '{term}': {e!s:.50}")

    unique = _dedup_by_name(results)
    print(f"    {len(unique)} sponsors from FDA applications")
    return unique


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE 6 — SEC EDGAR (US IPO filings)
# https://efts.sec.gov  — free, no key
# ══════════════════════════════════════════════════════════════════════════════

def fetch_edgar(sector="all", since_year=2015):
    print(f"\n[6/6] SEC EDGAR — recent oncology S-1 / IPO filings...")
    terms  = SECTOR_KEYWORDS.get(sector, SECTOR_KEYWORDS["all"])
    query  = " OR ".join(f'"{t}"' for t in terms[:3])
    url    = "https://efts.sec.gov/LATEST/search-index"
    params = {
        "q":        query,
        "dateRange":"custom",
        "startdt":  f"{since_year}-01-01",
        "enddt":    date.today().isoformat(),
        "forms":    "S-1,S-1/A",
    }
    try:
        r = requests.get(url, params=params, timeout=20,
                         headers={"User-Agent":"OncoSignal research@oncosignal.io"})
        r.raise_for_status()
        hits = r.json().get("hits",{}).get("hits",[])
        result = []
        seen   = set()
        for h in hits:
            src  = h.get("_source",{})
            name = src.get("entity_name","").strip()
            if not name or name.lower() in seen or _is_large(name):
                continue
            seen.add(name.lower())
            result.append({
                "name":         name,
                "source":       "SEC EDGAR (IPO/S-1)",
                "region":       "USA",
                "phases":       ["IPO Filed"],
                "conditions":   ["Oncology"],
                "trials":       [src.get("form_type","S-1")],
                "latest_start": _clean_date(src.get("file_date","")),
            })
        print(f"    {len(result)} companies from EDGAR")
        return result
    except Exception as e:
        print(f"    ERROR: {e}")
        return []


# ══════════════════════════════════════════════════════════════════════════════
# SHARED UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

def _is_large(name):
    n = name.lower()
    return any(ex in n for ex in EXCLUDE_ORGS)

def _clean_date(raw):
    """Return YYYY-MM-DD only if year is 2018-2026, else blank."""
    if not raw:
        return ""
    m = re.search(r"(20(?:1[89]|2[0-6]))-?(\d{2})?-?(\d{2})?", str(raw))
    if not m:
        return ""
    yr  = m.group(1)
    mon = m.group(2) or "01"
    day = m.group(3) or "01"
    return f"{yr}-{mon}-{day}"

def _extract_phase(text):
    t = text.lower()
    phases = []
    if re.search(r"phase\s*1|phase\s*i\b|phase i/ii", t):   phases.append("Phase 1")
    if re.search(r"phase\s*2|phase\s*ii\b", t):              phases.append("Phase 2")
    if re.search(r"phase\s*3|phase\s*iii\b", t):             phases.append("Phase 3")
    if re.search(r"pre.?clinical|preclinical|\bind\b", t):   phases.append("Pre-clinical")
    return phases

def _extract_conditions(text):
    types = [
        "lung cancer","breast cancer","prostate cancer","colorectal cancer",
        "ovarian cancer","pancreatic cancer","leukemia","lymphoma","melanoma",
        "glioblastoma","bladder cancer","kidney cancer","liver cancer",
        "gastric cancer","sarcoma","myeloma","nsclc","sclc","hcc",
        "neuroendocrine","thyroid cancer","cervical cancer",
    ]
    t = text.lower()
    return [ct.title() for ct in types if ct in t][:3]

def _dedup_by_name(companies):
    seen, result = {}, []
    for c in companies:
        key = re.sub(r"[^a-z0-9]", "", c["name"].lower())
        matched = next((k for k in seen if len(key) > 5 and (key in k or k in key)), None)
        if matched:
            seen[matched]["trials"].extend(c.get("trials",[]))
            seen[matched]["phases"].extend(c.get("phases",[]))
            seen[matched]["source"] += f" + {c.get('source','')}"
        else:
            seen[key] = c
            result.append(c)
    return result


# ══════════════════════════════════════════════════════════════════════════════
# SCORING ENGINE
# ══════════════════════════════════════════════════════════════════════════════

def score(c, sector):
    pts, signals = 0, []
    phases  = " ".join(c.get("phases",[])).lower()
    source  = c.get("source","")
    start   = c.get("latest_start","")
    n       = len(c.get("trials",[]))
    txt     = (phases + " " +
               " ".join(c.get("conditions",[])) + " " +
               " ".join(c.get("keywords",[]  ))).lower()

    # Stage
    if re.search(r"phase 1|phase i|pre.?clinical|ind\b", phases):
        pts += 30; signals.append("Phase I / Pre-clinical — no legacy data vendor")
    elif re.search(r"phase 2|phase ii", phases):
        pts += 20; signals.append("Phase II — building medical affairs stack")
    elif re.search(r"phase 3|phase iii", phases):
        pts += 10; signals.append("Phase III active program")
    elif "nda" in phases or "bla" in phases or "ipo" in phases:
        pts += 12; signals.append("FDA application / IPO stage")
    else:
        pts += 15; signals.append("Active oncology program")

    # Recency (only valid years)
    if start:
        try:
            yr = int(start[:4])
            if 2018 <= yr <= 2026:
                if yr >= 2023:   pts += 25; signals.append(f"Active since {yr} — very new entrant")
                elif yr >= 2021: pts += 15; signals.append(f"Active since {yr}")
                elif yr >= 2019: pts +=  8
        except ValueError:
            pass

    # Sector keyword match
    matched = [t for t in SECTOR_KEYWORDS.get(sector,[]) if t.lower() in txt]
    if matched:
        pts += 20; signals.append(f"Sector match: {matched[0]}")

    # Source bonuses
    if "WHO"         in source: pts += 10; signals.append("Global registry — non-US company")
    if "EDGAR"       in source: pts += 12; signals.append("Recently filed S-1 / IPO")
    if "Preprint"    in source: pts +=  8; signals.append("Publishing preprints — pre-IND stage")
    if "PubMed"      in source: pts +=  5; signals.append("Active research publications")
    if "Google News" in source:
        pts += 20  # News = earliest possible signal for pharma companies
        funding = c.get("funding", "")
        series  = c.get("series",  "")
        nsig    = c.get("news_signal", "")
        if series in ("Series A", "Seed"):
            pts += 15; signals.append(f"Just raised {series} {funding} — no data vendors yet".strip())
        elif series == "Series B":
            pts += 10; signals.append(f"Series B {funding} — building infrastructure now".strip())
        elif "stealth" in nsig or "launch" in nsig:
            pts += 12; signals.append(f"Emerged from stealth / newly launched {funding}".strip())
        elif funding:
            pts +=  8; signals.append(f"Funding news: {funding} {series}".strip())

    # Focused pipeline
    if 1 <= n <= 3: pts += 10; signals.append(f"{n} focused program(s)")

    pts = min(pts, 100)
    pri = "HOT" if pts >= 65 else "WARM" if pts >= 40 else "MONITOR"

    c["score"]         = pts
    c["priority"]      = pri
    c["why"]           = " | ".join(signals[:3])
    c["phase_display"] = ", ".join(dict.fromkeys(c.get("phases",[]))) or "Pre-clinical / Early"
    c["cond_display"]  = ", ".join(dict.fromkeys(c.get("conditions",[])))[:60] or "Oncology"
    return c


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT  — same style as oncology_prospects.xlsx
# ══════════════════════════════════════════════════════════════════════════════

C = {
    "purple_dark":"3B1F8C","purple_mid":"5C3DC7","purple_pale":"F5F3FF",
    "hot_bg":"FFF0EE","hot_fg":"C0392B",
    "warm_bg":"FFF8EC","warm_fg":"9A6200",
    "mon_bg":"EDFAF3","mon_fg":"1A7A4A",
    "border":"D5CCF5","grey_row":"F8F8F8","white":"FFFFFF",
}

def _fill(h): return PatternFill("solid", fgColor=h)

def _border(top=False):
    t = Side(style="thin",   color=C["border"])
    k = Side(style="medium", color=C["purple_mid"])
    return Border(left=t, right=t, top=k if top else t, bottom=t)

def write_excel(companies, sector, since_year, sources_used, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Oncology Prospects"

    # Row 1 — Title
    ws.merge_cells("A1:K1")
    t = ws["A1"]
    t.value = "OncoSignal — Global Early-Stage Oncology Prospect List"
    t.font  = Font(name="Arial", size=14, bold=True, color=C["white"])
    t.fill  = _fill(C["purple_dark"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Row 2 — Subtitle
    ws.merge_cells("A2:K2")
    s = ws["A2"]
    s.value = (f"Sources: {sources_used}  |  Sector: {sector}  |  "
               f"Since: {since_year}  |  Run: {datetime.now().strftime('%d %b %Y %H:%M')}")
    s.font  = Font(name="Arial", size=9, italic=True, color=C["purple_mid"])
    s.fill  = _fill(C["purple_pale"])
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # Row 3 — Headers
    headers = ["#","Company Name","Region","Phase",
               "Cancer Type / Condition",
               "Key Signal — Why They Need Your Platform",
               "Priority","Score","Latest Activity","Source","Verify Link"]
    widths  = [4,  30,           12,     16,
               26,
               50,
               10,      8,    16,              22,      32]

    ws.row_dimensions[3].height = 22
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font      = Font(name="Arial", size=10, bold=True, color=C["white"])
        cell.fill      = _fill(C["purple_mid"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _border(top=True)
        ws.column_dimensions[get_column_letter(col)].width = w

    # Data rows
    for i, c in enumerate(companies, 1):
        row    = 3 + i
        ws.row_dimensions[row].height = 40
        pri    = c.get("priority","MONITOR")
        pri_bg = C["hot_bg"]   if pri=="HOT"  else C["warm_bg"] if pri=="WARM"  else C["mon_bg"]
        pri_fg = C["hot_fg"]   if pri=="HOT"  else C["warm_fg"] if pri=="WARM"  else C["mon_fg"]
        row_bg = C["grey_row"] if i % 2 == 0  else C["white"]
        name   = c.get("name","")
        region = c.get("region","")

        # Verify link: WHO for non-US, ClinicalTrials for US
        if region and region not in ("USA",""):
            verify = f"https://trialsearch.who.int/?TrialID=&Title=&PrimarySponsors={requests.utils.quote(name)}"
        else:
            verify = f"https://clinicaltrials.gov/search?spons={requests.utils.quote(name)}"

        row_data = [
            i, name, region,
            c.get("phase_display",""),
            c.get("cond_display",""),
            c.get("why",""),
            pri,
            c.get("score",0),
            c.get("latest_start","")[:10] or "—",
            c.get("source","")[:28],
            verify,
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = _border()
            if col == 7:    # Priority
                cell.fill      = _fill(pri_bg)
                cell.font      = Font(name="Arial", size=10, bold=True, color=pri_fg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 8:  # Score
                cell.fill      = _fill(row_bg)
                cell.font      = Font(name="Arial", size=10, bold=True, color=C["purple_mid"])
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 1:  # Row num
                cell.fill      = _fill(row_bg)
                cell.font      = Font(name="Arial", size=9, color="888888")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 2:  # Company name
                cell.fill      = _fill(row_bg)
                cell.font      = Font(name="Arial", size=10, bold=True, color="1A1A2E")
                cell.alignment = Alignment(vertical="center")
            elif col in (6, 11):  # Why / Verify — wrap
                cell.fill      = _fill(row_bg)
                cell.font      = Font(name="Arial", size=9, color="444444")
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            else:
                cell.fill      = _fill(row_bg)
                cell.font      = Font(name="Arial", size=9, color="333333")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Summary row
    last = 3 + len(companies)
    sr   = last + 2
    ws.cell(row=sr, column=1, value="TOTAL").font = Font(name="Arial", size=9, bold=True, color=C["purple_dark"])
    ws.cell(row=sr, column=2,
            value=f'=COUNTIF(G4:G{last},"HOT")&" HOT  |  "&COUNTIF(G4:G{last},"WARM")&" WARM  |  "&COUNTIF(G4:G{last},"MONITOR")&" MONITOR  |  "&COUNTA(B4:B{last})&" total companies"')
    ws.cell(row=sr, column=2).font = Font(name="Arial", size=9, color="555555")

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:K{last}"
    ws.sheet_properties.tabColor = C["purple_mid"]
    wb.save(out_path)
    print(f"    Saved → {out_path}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
# SOURCE 7 — Google News RSS  (free, no key, no scraping — official RSS feed)
# Catches pharma companies at the EARLIEST signal: funding announcement,
# stealth launch, partnership deal — before any trial is filed.
# Query format: https://news.google.com/rss/search?q=QUERY&hl=en-US&gl=US&ceid=US:en
# ══════════════════════════════════════════════════════════════════════════════

# Search queries tuned to catch NEW pharma company signals
GNEWS_QUERIES = {
    "radiopharmaceutical": [
        "radiopharmaceutical oncology raises OR launch OR series",
        "theranostics biotech raises OR founded OR series",
        "radioligand therapy company series A OR series B",
    ],
    "immuno-oncology": [
        "immunotherapy oncology series A OR series B raises",
        "checkpoint inhibitor biotech launch OR founded OR raises",
        "CAR-T company series A OR series B OR raises",
    ],
    "adc": [
        "antibody drug conjugate ADC biotech raises OR series",
        "ADC oncology company launch OR series A OR series B",
    ],
    "precision": [
        "precision oncology biotech raises series A OR series B",
        "targeted therapy company launch OR founded OR raises",
    ],
    "cell therapy": [
        "cell therapy oncology series A OR series B raises",
        "CAR-T NK cell company raises OR launch OR founded",
    ],
    "rna": [
        "mRNA oncology company raises OR series A OR founded",
        "RNA therapeutics cancer series A OR series B",
    ],
    "all": [
        "oncology biotech raises series A OR series B 2023 OR 2024",
        "cancer therapeutics company launch OR founded raises",
        "pharma oncology series A raises OR stealth OR launch",
        "radiopharmaceutical theranostics raises OR founded OR series",
    ],
}

# Signals that indicate a NEW, SMALL company (high prospect value)
FUNDING_SIGNALS  = ["series a","series b","seed round","raises","raised","stealth","launch","founded","emerges","new company","ipo"]
EXCLUDE_SIGNALS  = ["acqui","merger","phase 3 results","approval","fda approved","partnership with pfizer","partnership with roche"]


def fetch_gnews(sector="all", since_year=2020):
    print(f"\n[7/7] Google News RSS — funding & launch announcements...")
    queries  = GNEWS_QUERIES.get(sector, GNEWS_QUERIES["all"])
    results  = []

    for query in queries:
        # Build Google News RSS URL
        encoded = requests.utils.quote(query)
        url     = f"https://news.google.com/rss/search?q={encoded}&hl=en-US&gl=US&ceid=US:en"
        try:
            r = requests.get(url, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            }, timeout=15)
            if r.status_code != 200:
                continue
            parsed = _parse_gnews_rss(r.text, query, since_year)
            results.extend(parsed)
            time.sleep(1.2)  # polite delay between queries
        except Exception as e:
            print(f"    Google News '{query[:40]}': {e!s:.60}")

    unique = _dedup_by_name(results)
    print(f"    {len(unique)} companies from Google News RSS")
    return unique


def _parse_gnews_rss(xml_text, search_term, since_year):
    companies = []
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        return []

    for item in root.iter("item"):
        title   = item.findtext("title",       "")
        desc    = item.findtext("description", "")
        pub_raw = item.findtext("pubDate",     "")
        full    = title + " " + desc

        # Date filter — only items from since_year onwards
        pub_date = _parse_pub_date(pub_raw)
        if pub_date:
            try:
                if int(pub_date[:4]) < since_year:
                    continue
            except ValueError:
                pass

        # Skip if it's about a large pharma company
        if any(ex in full.lower() for ex in EXCLUDE_ORGS):
            continue

        # Skip acquisition/approval news (not new companies)
        if any(sig in full.lower() for sig in EXCLUDE_SIGNALS):
            continue

        # Must contain a funding/launch signal to be relevant
        has_signal = any(sig in full.lower() for sig in FUNDING_SIGNALS)
        if not has_signal:
            continue

        # Extract company name from headline
        name = _extract_company_from_headline(title)
        if not name or _is_large(name) or len(name) < 4:
            continue

        # Extract funding amount & series if mentioned
        funding = _extract_funding(full)
        series  = _extract_series(full)

        # Build signal description
        signal_parts = []
        if funding: signal_parts.append(funding)
        if series:  signal_parts.append(series)
        if "stealth" in full.lower() or "emerges" in full.lower():
            signal_parts.append("emerged from stealth")
        if "founded" in full.lower() or "launch" in full.lower():
            signal_parts.append("newly launched")
        signal_str = " · ".join(signal_parts) if signal_parts else "funding/launch news"

        companies.append({
            "name":         name,
            "source":       "Google News RSS",
            "region":       _extract_region(full),
            "phases":       _extract_phase(full),
            "conditions":   _extract_conditions(full),
            "trials":       [title[:100]],
            "latest_start": pub_date,
            "funding":      funding,
            "series":       series,
            "news_signal":  signal_str,
            "headline":     title[:120],
        })

    return companies


def _extract_company_from_headline(title):
    """Extract company name from headlines like:
    'Ratio Therapeutics Raises $90M Series B...'
    'Abdera Therapeutics Emerges from Stealth with $142M...'
    """
    # Primary: match Company Name + action verb
    m = re.match(
        r"^([A-Z][A-Za-z0-9\-]+(?: [A-Z][A-Za-z0-9\-]+){0,4})\s+"
        r"(?:Raises?|Announces?|Launches?|Emerges?|Closes?|Secures?|"
        r"Receives?|Completes?|Reports?|Names?|Appoints?|Files?|"
        r"Starts?|Initiates?|Presents?|Wins?|Gets?|Doses?|Enters?|"
        r"Partners?|Licenses?|Acquires?|Expands?|Begins?)",
        title.strip()
    )
    if m:
        name = m.group(1).strip()
        if len(name) > 3 and name.lower() not in ("the","a","an","new","study","trial","data","fda","nih"):
            return name

    # Fallback: first 1-4 words containing biotech suffix
    suffixes = ["therapeutics","pharmaceuticals","biosciences","bioscience",
                "biotech","biologics","oncology","medicines","biopharma","bio"]
    words = title.split()
    for i in range(1, min(6, len(words))):
        chunk = " ".join(words[:i])
        if any(s in chunk.lower() for s in suffixes):
            return chunk
    return ""


def _extract_funding(text):
    m = re.search(r"\$([\d\.]+)\s*(M|B|million|billion)\b", text, re.I)
    if m:
        amt  = m.group(1)
        unit = "M" if m.group(2).upper().startswith("M") else "B"
        return f"${amt}{unit}"
    return ""


def _extract_series(text):
    m = re.search(r"\bSeries\s+([A-D])\b", text, re.I)
    return f"Series {m.group(1).upper()}" if m else ""


def _extract_region(text):
    region_map = {
        "Australia":"Australia","ASX":"Australia","ANZCTR":"Australia",
        "United Kingdom":"UK","London":"UK","LSE":"UK",
        "Germany":"Germany","France":"France","Switzerland":"Switzerland",
        "Canada":"Canada","TSX":"Canada","Toronto":"Canada",
        "Japan":"Japan","China":"China","South Korea":"South Korea",
        "Israel":"Israel","Netherlands":"Netherlands","Sweden":"Sweden",
        "San Francisco":"USA","Boston":"USA","New York":"USA",
        "Cambridge, MA":"USA","Nasdaq":"USA","NYSE":"USA",
    }
    for k, v in region_map.items():
        if k in text:
            return v
    return ""


def _parse_pub_date(raw):
    """Parse RSS pubDate like 'Mon, 15 Jan 2024 10:00:00 GMT'"""
    if not raw:
        return ""
    # Try RFC 822 format
    for fmt in ("%a, %d %b %Y %H:%M:%S %Z", "%a, %d %b %Y %H:%M:%S %z",
                "%a, %d %b %Y"):
        try:
            dt = datetime.strptime(raw[:25].strip(), fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Fallback: extract year
    m = re.search(r"(20(?:1[89]|2[0-6]))", raw)
    return f"{m.group(1)}-01-01" if m else ""


SOURCE_MAP = {
    "ct":      ("ClinicalTrials.gov",  fetch_clinicaltrials),
    "who":     ("WHO ICTRP",           fetch_who_ictrp),
    "pubmed":  ("PubMed/NCBI",         fetch_pubmed),
    "biorxiv": ("bioRxiv/medRxiv",     fetch_biorxiv),
    "openfda": ("OpenFDA",             fetch_openfda),
    "edgar":   ("SEC EDGAR",           fetch_edgar),
    "gnews":   ("Google News RSS",     fetch_gnews),
}

def main():
    parser = argparse.ArgumentParser(
        description="OncoSignal v2 — global oncology prospects from free APIs",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python oncology_pipeline_v2.py
  python oncology_pipeline_v2.py --sector radiopharmaceutical
  python oncology_pipeline_v2.py --sector all --since 2021 --count 40
  python oncology_pipeline_v2.py --sources who,pubmed,biorxiv
  python oncology_pipeline_v2.py --sources ct,who,edgar --sector immuno-oncology
  python oncology_pipeline_v2.py --sources gnews --sector radiopharmaceutical
  python oncology_pipeline_v2.py --sources gnews,who,ct --sector all
        """
    )
    parser.add_argument("--sector",  default="all",  choices=list(SECTOR_KEYWORDS.keys()))
    parser.add_argument("--since",   default=2020,   type=int,  help="Active since year")
    parser.add_argument("--count",   default=30,     type=int,  help="Max companies in output")
    parser.add_argument("--sources", default="ct,who,pubmed,biorxiv,openfda,edgar,gnews",
                        help="Comma-separated: ct,who,pubmed,biorxiv,openfda,edgar,gnews")
    parser.add_argument("--output",  default="",     help="Output filename (auto if blank)")
    args    = parser.parse_args()
    sources = [s.strip().lower() for s in args.sources.split(",")]

    print("=" * 58)
    print("  OncoSignal Free Pipeline v2")
    print(f"  Sector  : {args.sector}")
    print(f"  Since   : {args.since}")
    print(f"  Sources : {', '.join(sources)}")
    print("  Why v2  : WHO ICTRP (global) + Google News RSS (funding signals)")
    print("=" * 58)

    all_companies  = []
    sources_labels = []

    for key in sources:
        if key not in SOURCE_MAP:
            print(f"  Unknown source '{key}' — skipping")
            continue
        label, fn = SOURCE_MAP[key]
        sources_labels.append(label)
        # Pass only args the function accepts
        if key in ("ct", "pubmed"):
            all_companies.extend(fn(args.sector, args.since, 100))
        elif key in ("who", "biorxiv", "openfda", "edgar", "gnews"):
            all_companies.extend(fn(args.sector, args.since))
        time.sleep(0.5)

    if not all_companies:
        print("\nNo data retrieved. Check your internet connection.")
        return

    print(f"\nRaw total          : {len(all_companies)}")
    all_companies = _dedup_by_name(all_companies)
    print(f"After deduplication: {len(all_companies)}")

    scored = sorted([score(c, args.sector) for c in all_companies],
                    key=lambda x: x["score"], reverse=True)[:args.count]

    hot  = sum(1 for c in scored if c["priority"]=="HOT")
    warm = sum(1 for c in scored if c["priority"]=="WARM")
    mon  = sum(1 for c in scored if c["priority"]=="MONITOR")

    print(f"\nResults : {len(scored)} companies  |  {hot} HOT  |  {warm} WARM  |  {mon} MONITOR")
    print("-" * 58)
    for c in scored[:12]:
        reg = f"[{c.get('region','')}]" if c.get("region") else ""
        print(f"  [{c['priority']:7}] {c['score']:3}/100  {c['name'][:32]:32} {reg}")
    if len(scored) > 12:
        print(f"  ... and {len(scored)-12} more in Excel")

    ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
    outfile = args.output or f"oncology_prospects_{args.sector}_{ts}.xlsx"
    print(f"\nWriting Excel...")
    write_excel(scored, args.sector, args.since, " · ".join(sources_labels), outfile)
    print("\nDone.")

if __name__ == "__main__":
    main()

# phase 1 and phase in breast cancer and neuroblastoma